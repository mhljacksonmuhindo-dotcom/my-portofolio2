import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QLineEdit, QFileDialog, QMessageBox, QLabel
from PyQt5.QtCore import Qt
import openpyxl
from openpyxl.styles import Font, PatternFill

class ExcelApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Interface QtPy5 + Excel')
        self.setGeometry(100, 100, 600, 400)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout = QVBoxLayout(central_widget)
        
        file_layout = QHBoxLayout()
        self.file_label = QLabel('Aucun fichier sélectionné')
        file_layout.addWidget(self.file_label)
        
        self.select_btn = QPushButton('Sélectionner Excel')
        self.select_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.select_btn)
        layout.addLayout(file_layout)
        
        
        input_layout = QHBoxLayout()
        self.range_edit = QLineEdit('A1:B10')
        self.range_edit.setPlaceholderText('Plage (ex: A1:B10)')
        input_layout.addWidget(QLabel('Plage:'))
        input_layout.addWidget(self.range_edit)
        
        self.value_edit = QLineEdit()
        self.value_edit.setPlaceholderText('Valeur à mettre')
        input_layout.addWidget(QLabel('Valeur:'))
        input_layout.addWidget(self.value_edit)
        
        layout.addLayout(input_layout)
        
        
        btn_layout = QHBoxLayout()
        self.write_btn = QPushButton('Écrire dans Excel')
        self.write_btn.clicked.connect(self.write_cell)
        btn_layout.addWidget(self.write_btn)
        
        self.format_btn = QPushButton('Formater plage')
        self.format_btn.clicked.connect(self.format_range)
        btn_layout.addWidget(self.format_btn)
        
        self.save_btn = QPushButton('Sauvegarder')
        self.save_btn.clicked.connect(self.save_file)
        btn_layout.addWidget(self.save_btn)
        
        layout.addLayout(btn_layout)
        
        self.wb = None
        self.ws = None
        
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Ouvrir Excel', '', 'Excel files (*.xlsx)')
        if file_path:
            self.file_label.setText(file_path.split('/')[-1])
            self.wb = openpyxl.load_workbook(file_path)
            self.ws = self.wb.active
    
    def write_cell(self):
        if not self.ws:
            QMessageBox.warning(self, 'Erreur', 'Sélectionnez un fichier Excel!')
            return
        range_str = self.range_edit.text()
        value = self.value_edit.text()
        try:
            cell = self.ws[range_str]
            cell.value = value
            QMessageBox.information(self, 'Succès', f'Valeur écrite dans {range_str}')
        except:
            QMessageBox.warning(self, 'Erreur', 'Plage invalide!')
    
    def format_range(self):
        if not self.ws:
            QMessageBox.warning(self, 'Erreur', 'Sélectionnez un fichier Excel!')
            return
        range_str = self.range_edit.text()
        try:
            for row in self.ws[range_str]:
                for cell in row:
                    cell.font = Font(bold=True, color='FF0000')
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            QMessageBox.information(self, 'Succès', f'Plage {range_str} formatée!')
        except:
            QMessageBox.warning(self, 'Erreur', 'Plage invalide!')
    
    def save_file(self):
        if not self.wb:
            QMessageBox.warning(self, 'Erreur', 'Aucun fichier ouvert!')
            return
        file_path, _ = QFileDialog.getSaveFileName(self, 'Sauvegarder Excel', '', 'Excel files (*.xlsx)')
        if file_path:
            self.wb.save(file_path)
            QMessageBox.information(self, 'Succès', 'Fichier sauvegardé!')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ExcelApp()
    window.show()
    sys.exit(app.exec_())

